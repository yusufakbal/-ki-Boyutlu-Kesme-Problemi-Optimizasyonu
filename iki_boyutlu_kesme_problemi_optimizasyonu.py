import xlrd
import random
from openpyxl import *



class sac_kesme():
    def __init__(self,stok_kodu=0, kesim_adet= 0,cins="cins", kalinlik = 0, en=0, boy = 0, sac_kod_listesi= [],kesim_adedi_listesi=[], sac_cinsi_listesi =[],sac_kalinlik_listesi= [],sac_en_listesi= [],sac_boy_listesi = [],
                 ana_sac_kodu_listesi = [],ana_sac_cinsi= [],ana_sac_kalinlik_listesi= [], ana_sac_en_listesi = [], ana_sac_boy_listesi= [],sknps_0_5_listesi=[], sknps_0_8_listesi=[], sknps_1_listesi=[], glvnz_0_7_listesi=[],
                 glvnz_1_5_listesi=[], glvnz_2_listesi=[], mkrom_0_5_listesi=[], mkrom_1_listesi=[],mkrom_1_5_listesi=[], mkrom_2_listesi=[], tkrom_0_8_listesi=[], tkrom_1_5_listesi=[],hrp_2_5_listesi =[] ,hrp_3_listesi = [] ,
                 hrp_10_listesi = [] ,dkp_2_5_listesi = [],ferforje_1_listesi = [],dkrom_0_8_listesi = [],index = 0, deger1 = 0, deger2 = 0, tum_sac_yerlesimleri = [], sacyerlesim = [], toplam_fire= []):
        self.stok_kodu = stok_kodu
        self.kesim_adet = kesim_adet
        self.cins = cins
        self.kalinlik = kalinlik
        self.en = en
        self.boy = boy
        self.sac_kod_listesi = sac_kod_listesi
        self.kesim_adedi_listesi = kesim_adedi_listesi
        self.sac_cinsi_listesi = sac_cinsi_listesi
        self.sac_kalinlik_listesi = sac_kalinlik_listesi
        self.sac_en_listesi = sac_en_listesi
        self.sac_boy_listesi = sac_boy_listesi
        self.ana_sac_kodu_listesi = ana_sac_kodu_listesi
        self.ana_sac_cinsi = ana_sac_cinsi
        self.ana_sac_kalinlik_listesi = ana_sac_kalinlik_listesi
        self.ana_sac_en_listesi = ana_sac_en_listesi
        self.ana_sac_boy_listesi = ana_sac_boy_listesi
        self.sknps_0_5_listesi = sknps_0_5_listesi
        self.sknps_0_8_listesi = sknps_0_8_listesi
        self.sknps_1_listesi = sknps_1_listesi
        self.glvnz_0_7_listesi = glvnz_0_7_listesi
        self.glvnz_1_5_listesi = glvnz_1_5_listesi
        self.glvnz_2_listesi = glvnz_2_listesi
        self.mkrom_0_5_listesi = mkrom_0_5_listesi
        self.mkrom_1_listesi = mkrom_1_listesi
        self.mkrom_1_5_listesi = mkrom_1_5_listesi
        self.mkrom_2_listesi = mkrom_2_listesi
        self.tkrom_0_8_listesi = tkrom_0_8_listesi
        self.tkrom_1_5_listesi = tkrom_1_5_listesi
        self.hrp_2_5_listesi = hrp_2_5_listesi
        self.hrp_3_listesi = hrp_3_listesi
        self.hrp_10_listesi = hrp_10_listesi
        self.dkp_2_5_listesi = dkp_2_5_listesi
        self.ferforje_1_listesi = ferforje_1_listesi
        self.dkrom_0_8_listesi = dkrom_0_8_listesi
        self.index = index
        self.deger1 = deger1
        self.deger2 = deger2
        self.tum_sac_yerlesimleri = tum_sac_yerlesimleri
        self.sacyerlesim = sacyerlesim
        self.toplam_fire = toplam_fire





    def veri_cek(self):
        file = "C:/Users/ASUS/Desktop/DATA.xls"
        rb = xlrd.open_workbook(file)
        sheet = rb.sheet_by_index(0)
        data = [[sheet.cell_value(r,c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        self.sac_kod_listesi= []
        self.kesim_adedi_listesi= []
        self.sac_cinsi_listesi= []
        self.sac_kalinlik_listesi = []
        self.sac_en_listesi= []
        self.sac_boy_listesi = []




        for i in range(1,sheet.nrows):
            self.stok_kodu = data[i][2]
            self.kesim_adet= data[i][4]
            self.kesim_adet = int(self.kesim_adet)
            self.cins = data[i][5]
            self.kalinlik = data[i][6]
            self.kalinlik = float(self.kalinlik)
            self.en = data[i][7]
            self.en = float(self.en)
            self.boy = data[i][8]
            self.boy = float(self.boy)
            for j in range(0, self.kesim_adet):

                self.sac_kod_listesi.append(self.stok_kodu)
                self.kesim_adedi_listesi.append(1)
                self.sac_cinsi_listesi.append(self.cins)
                self.sac_kalinlik_listesi.append(self.kalinlik)
                self.sac_en_listesi.append(self.en)
                self.sac_boy_listesi.append(self.boy)

        file = "C:/Users/ASUS/Desktop/SAC.xlsx"
        rb = xlrd.open_workbook(file)
        sheet = rb.sheet_by_index(0)
        data = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        for i in range(1,sheet.nrows):
            self.stok_kodu = data[i][0]
            self.cins = data[i][1]
            self.kalinlik = data[i][3]
            self.kalinlik = float(self.kalinlik)
            self.en = data[i][4]
            self.en = float(self.en)
            self.boy = data[i][5]
            self.boy = float(self.boy)
            self.ana_sac_kodu_listesi.append(self.stok_kodu)
            self.ana_sac_cinsi.append(self.cins)
            self.ana_sac_kalinlik_listesi.append(self.kalinlik)
            self.ana_sac_en_listesi.append(self.en)
            self.ana_sac_boy_listesi.append(self.boy)



        return self.sac_kod_listesi, self.kesim_adedi_listesi, self.sac_cinsi_listesi,\
               self.sac_kalinlik_listesi, self.sac_en_listesi, self.sac_boy_listesi,\
               self.ana_sac_boy_listesi, self.ana_sac_en_listesi, self.ana_sac_kalinlik_listesi, self.ana_sac_cinsi,self.ana_sac_kodu_listesi

    def grupla(self):

        while (self.index < len(self.sac_cinsi_listesi)):

            self.deger1 = self.sac_cinsi_listesi[self.index]
            self.deger2 = self.sac_kalinlik_listesi[self.index]

            if (self.deger1 == "M.KROM" and self.deger2 == 0.5) :
                self.mkrom_0_5_listesi.append(self.index)

            elif (self.deger1 == "M.KROM" and self.deger2 == 1) :
                self.mkrom_1_listesi.append(self.index)

            elif (self.deger1 == "M.KROM" and self.deger2 == 1.5) :
                self.mkrom_1_5_listesi.append(self.index)

            elif (self.deger1 == "M.KROM" and self.deger2 == 2):
                self.mkrom_2_listesi.append(self.index)

            elif (self.deger1 == "SKNPS" and self.deger2 == 0.5) :
                self.sknps_0_5_listesi.append(self.index)

            elif (self.deger1 == "SKNPS" and self.deger2 == 0.8) :
                self.sknps_0_8_listesi.append(self.index)

            elif (self.deger1 == "SKNPS" and self.deger2 == 1) :
                self.sknps_1_listesi.append(self.index)

            elif (self.deger1 == "GLVNZ" and self.deger2 == 0.7) :
                self.glvnz_0_7_listesi.append(self.index)

            elif (self.deger1 == "GLVNZ" and self.deger2 == 1.5) :
                self.glvnz_1_5_listesi.append(self.index)

            elif (self.deger1 == "GLVNZ" and self.deger2 == 2) :
                self.glvnz_2_listesi.append(self.index)

            elif (self.deger1 == "T.KROM" and self.deger2 == 0.8) :
                self.tkrom_0_8_listesi.append(self.index)

            elif (self.deger1 == "T.KROM" and self.deger2 == 1.5) :
                self.tkrom_1_5_listesi.append(self.index)

            elif (self.deger1 == "HRP" and self.deger2 == 2.5) :
                self.hrp_2_5_listesi.append(self.index)

            elif (self.deger1 == "HRP" and self.deger2 == 3) :
                self.hrp_3_listesi.append(self.index)

            elif (self.deger1 == "HRP" and self.deger2 == 10) :
                self.hrp_10_listesi.append(self.index)

            elif (self.deger1 == "DKP" and self.deger2 == 2.5) :
                self.dkp_2_5_listesi.append(self.index)

            elif (self.deger1 == "FERFORJE" and self.deger2 == 1) :
                self.ferforje_1_listesi.append(self.index)

            elif (self.deger1 == "D.KROM" and self.deger2 == 0.8) :
                self.dkrom_0_8_listesi.append(self.index)

            else:
                print(self.deger1,self.deger2, "Böyle bir grup bulunamamıştır...")

            self.index += 1

        return self.tkrom_0_8_listesi, self.sknps_0_8_listesi, self.sknps_1_listesi, self.glvnz_0_7_listesi,\
               self.glvnz_1_5_listesi, self.glvnz_2_listesi, self.mkrom_0_5_listesi, self.mkrom_1_listesi,\
               self.mkrom_1_5_listesi, self.mkrom_2_listesi, self.sknps_0_5_listesi, self.tkrom_1_5_listesi,\
               self.hrp_2_5_listesi , self.hrp_3_listesi,self.hrp_10_listesi ,self.dkp_2_5_listesi ,self.ferforje_1_listesi, self.dkrom_0_8_listesi

    def uygunluk(self, liste):

        print(liste)
        i = liste[0]

        self.tum_sac_yerlesimleri = []
        self.toplam_fire = []
        anasaclar = []
        anasac_indis = 0

        for c, k, e, b in zip(self.ana_sac_cinsi, self.ana_sac_kalinlik_listesi, self.ana_sac_en_listesi,
                              self.ana_sac_boy_listesi):
            parcalar = []
            for a in liste:
                parcalar.append(a)


            if (self.sac_cinsi_listesi[i] == c and self.sac_kalinlik_listesi[i] == k):

                print("***********ana sac ölçüsü*************")
                print("******cins: {} , kalınlık: {}  ,  en:{}  , boy: {}******".format(c,k,e,b))
                print("")



                x = 0
                y = 0
                y1 = e
                x1 = b
                yenb = self.sac_en_listesi[i]


                self.sacyerlesim = []
                bosliste = []
                self.sacyerlesim.append(bosliste)
                index = 0
                kosul = 0

                while True:
                    z=0


                    while (z< len(parcalar)):

                        gen = parcalar[z]
                        y2 = self.sac_en_listesi[gen]
                        x2 = self.sac_boy_listesi[gen]
                        print(gen, ".index değeri en:{}, boy:{}".format(y2,x2))
                        z +=1




                        if (x1 - x - x2 >= 0 and y1 - y2 >= 0):
                            self.sacyerlesim[index].append(gen)
                            parcalar.remove(gen)
                            y1 = y2
                            x = x + x2
                            print("kabul edilen gen değeri:", gen )
                            print("")
                            z-=1




                    for gen in parcalar:
                        if len(self.sacyerlesim[index]) == 0 :

                            break

                        y2 = self.sac_en_listesi[gen]

                        if (e - yenb - y2 >= 0):
                            bosliste = []
                            self.sacyerlesim.append(bosliste)
                            index += 1
                            y = yenb
                            y1 = e - yenb
                            x = 0
                            x1 = b
                            yenb = yenb + self.sac_en_listesi[gen]
                            print("y bloğu kontrolü yapılıyor..")
                            print("")
                            break

                    if (index == kosul):
                        break
                    else:
                        kosul = index

                if(len(self.sacyerlesim[0]) != 0):
                    self.tum_sac_yerlesimleri.append(self.sacyerlesim)
                    print(self.tum_sac_yerlesimleri)

                alan = 0
                for p in self.sacyerlesim:
                    for f in p:
                        en1 = self.sac_en_listesi[f]
                        boy1 = self.sac_boy_listesi[f]
                        alan += ( en1 * boy1 )

                print("sac yerleşim : ",self.sacyerlesim)
                if(len(self.sacyerlesim[0]) != 0):
                    sac_alani = e * b
                    fire = sac_alani - alan
                    self.toplam_fire.append(fire)
                    anasaclar.append(anasac_indis)


            anasac_indis += 1

            if len(parcalar) == 0:
                break

        en_kucuk =100000000000000000
        ind = 0
        for i, j in enumerate(self.toplam_fire):

            if j < en_kucuk:
                en_kucuk = j
                ind = i

        print("anasacın indis değeri :",ind)

        en_kucuk_fire = self.toplam_fire[ind]
        en_kucuk_sac_yerlesimi = self.tum_sac_yerlesimleri[ind]
        en_kucuk_anasac = anasaclar[ind]


        return en_kucuk_sac_yerlesimi, en_kucuk_fire, en_kucuk_anasac

    def siralama(self):

        enler = self.sac_en_listesi
        siralanmis_en_listesi = []
        siralanmis_ind_listesi = []

        while True:
            en_buyuk = 0
            ind = 0
            for i,j in enumerate(enler):
                k = 0
                for b in siralanmis_ind_listesi:
                    if b == i:
                        k += 1
                if j>en_buyuk and k == 0:
                    en_buyuk = j
                    ind= i

            siralanmis_en_listesi.append(en_buyuk)
            siralanmis_ind_listesi.append(ind)

            if len(siralanmis_ind_listesi) == len(enler) :
                break
        self.sac_en_listesi = siralanmis_en_listesi

        duzenlenecek_listeler = [self.sac_kod_listesi,self.kesim_adedi_listesi,self.sac_cinsi_listesi,self.sac_kalinlik_listesi,self.sac_boy_listesi]
        g=0
        for i in duzenlenecek_listeler:

            a = 0
            bosliste = []
            while True:
                for k,l in enumerate(i):
                    if k == siralanmis_ind_listesi[a]:
                        bosliste.append(l)
                        break

                a+=1
                if len(i) == len(bosliste):
                    break
            duzenlenecek_listeler[g] = bosliste
            g+=1

        self.sac_kod_listesi = duzenlenecek_listeler[0]
        self.kesim_adedi_listesi = duzenlenecek_listeler[1]
        self.sac_cinsi_listesi = duzenlenecek_listeler[2]
        self.sac_kalinlik_listesi = duzenlenecek_listeler[3]
        self.sac_boy_listesi = duzenlenecek_listeler[4]

        return siralanmis_ind_listesi,self.sac_en_listesi, duzenlenecek_listeler,self.sac_kod_listesi,\
               self.kesim_adedi_listesi,self.sac_cinsi_listesi,self.sac_kalinlik_listesi,self.sac_boy_listesi





sac = sac_kesme()
veri = sac.veri_cek()
siralanmis = sac.siralama()
grup = sac.grupla()



tum_kromozomlar = []
tum_fireler = []
tum_anasac_bilgileri = []

for i in grup:

    if(len(i) != 0):
        kromozomlar = []
        anasac_bilgileri = []
        fireler = []
        f = 0
        while True:
            uygunluk_degeri = sac.uygunluk(i)
            kromozomlar.append(uygunluk_degeri[0])
            fireler.append(uygunluk_degeri[1])
            anasac_bilgileri.append(uygunluk_degeri[2])
            liste = kromozomlar[f]
            for j in liste:
                for r in j:
                    i.remove(r)

            f +=1
            if len(i) == 0:
                break


        toplam_fire = 0
        for fire in fireler:
            toplam_fire+= fire


        tum_kromozomlar.append(kromozomlar)
        tum_fireler.append(toplam_fire)
        tum_anasac_bilgileri.append(anasac_bilgileri)





saccins = veri[9]
sackalinlik = veri[8]
sacen = veri[7]
sacboy = veri[6]
sackod = veri[10]

print("saccins: {}\nsackalinlik: {}\nsacen: {}\nsacboy: {}".format(saccins,sackalinlik,sacen,sacboy))


for son_kromozom,en_kucuk,son_anasac in zip(tum_kromozomlar,tum_fireler,tum_anasac_bilgileri):



    print("Kromozom : {}\nfire : {}".format(son_kromozom,en_kucuk))

    print("anasac bilgileri :",son_anasac)

    for sac in son_anasac:
        cins = saccins[sac]
        kalinlik = sackalinlik[sac]
        en = sacen[sac]
        boy = sacboy[sac]


        print("Sac Cinsi: {}  Kalınlık: {}  En: {}  Boy: {}".format(cins,kalinlik,en,boy))



    enler = siralanmis[1]
    boylar = siralanmis[7]
    cinsi = siralanmis[5]
    kalinligi = siralanmis[6]
    kodu = siralanmis[3]

    for i in son_kromozom:
        for j in i:
            if(len(j) != 0):
                for z in j:
                    en = enler[z]
                    boy = boylar[z]
                    cins = cinsi[z]
                    kod = kodu[z]
                    kalinlik = kalinligi[z]
                    print(z,". deger ---- Kod: {}  Cins: {}  Kalınlık: {}  En: {}  Boy: {}".format(kod,cins,kalinlik,en,boy))





kitap = Workbook()
sheet = kitap.active


sheet.cell(row = 1 , column=1, value="PARÇA CİNSİ")
sheet.cell(row = 1 , column=2, value="ANASAC KODU")
sheet.cell(row = 1 , column=3, value="ANASAC ÖLÇÜLERİ")
sheet.cell(row = 1 , column=6, value="Y BLOĞU")
sheet.cell(row = 1 , column=7, value="X BLOĞU")
sheet.cell(row = 1 , column=8, value="PARÇA KODLARI")
sheet.cell(row = 1 , column=9, value="PARÇA ÖLÇÜLERİ")
sheet.cell(row = 1 , column=12, value="FİRELER")
sheet.cell(row = 1 , column=13, value="FİRE ORANI (%)")
sheet.cell(row = 1 , column=14, value="TOPLAM FİRE")
sheet.cell(row = 1 , column=15, value="TOPLAM FİRE ORANI (%)")
sheet.cell(row = 2 , column=3, value="KALINLIK")
sheet.cell(row = 2 , column=4, value="EN")
sheet.cell(row = 2 , column=5, value="BOY")
sheet.cell(row = 2 , column=9, value="KALINLIK")
sheet.cell(row = 2 , column=10, value="EN")
sheet.cell(row = 2 , column=11, value="BOY")




enler = siralanmis[1]
boylar = siralanmis[7]
cinsi = siralanmis[5]
kalinligi = siralanmis[6]
kodu = siralanmis[3]


ind = 3
s = 3
tum_sac_alan = 0
toplam_fireler = 0

for son_kromozom,en_kucuk,son_anasac in zip(tum_kromozomlar,tum_fireler,tum_anasac_bilgileri):

    toplam_sac_alan = 0

    for sac,i in zip(son_anasac,son_kromozom):


        cins = saccins[sac]
        kalinlik = sackalinlik[sac]
        en = sacen[sac]
        boy = sacboy[sac]
        kod = sackod[sac]

        sacalan = en * boy

        sheet.cell(row=ind, column=2, value=kod)
        sheet.cell(row=ind, column=3, value=kalinlik)
        sheet.cell(row=ind, column=4, value=en)
        sheet.cell(row=ind, column=5, value=boy)
        kalinlik = str(kalinlik)
        sheet.cell(row=ind, column=1, value=cins + " " + kalinlik)

        yblok = 1
        yblok = str(yblok)
        parcaalan = 0
        for j in i:
            if (len(j) != 0):

                y_blogu = "Y" + yblok
                sheet.cell(row=ind, column=6, value=y_blogu)
                yblok = int(yblok)
                yblok += 1
                yblok = str(yblok)


                xblok = 1
                xblok = str(xblok)

                for z in j:
                    x_blogu = "X" + xblok
                    sheet.cell(row=ind, column=7, value=x_blogu)
                    xblok = int(xblok)
                    xblok += 1
                    xblok = str(xblok)

                    en = enler[z]
                    boy = boylar[z]
                    kod = kodu[z]
                    kalinlik = kalinligi[z]

                    sheet.cell(row=ind, column=8, value=kod)
                    sheet.cell(row=ind, column=9, value=kalinlik)
                    sheet.cell(row=ind, column=10, value=en)
                    sheet.cell(row=ind, column=11, value=boy)

                    parcaalan += en * boy

                    ind += 1
                    s += 1


        toplam_sac_alan += sacalan
        tum_sac_alan += sacalan
        fire = sacalan - parcaalan
        fireorani = fire*100/sacalan
        sheet.cell(row=ind - 1, column=12, value=fire)
        sheet.cell(row=ind - 1, column=13, value=fireorani)


    toplam_fire_orani = en_kucuk*100/toplam_sac_alan
    sheet.cell(row=s - 1, column=15, value=toplam_fire_orani)
    sheet.cell(row=s - 1, column=14, value=en_kucuk)

    toplam_fireler += en_kucuk

tum_fire_orani = toplam_fireler*100/tum_sac_alan
sheet.cell(row=ind + 3, column=13, value="Toplam Fire:")
sheet.cell(row=ind + 3, column=14, value=tum_fire_orani)




kitap.save("C:/Users/ASUS/Desktop/deneme_dosyası.xlsx")
kitap.close()